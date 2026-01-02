from typing import Any, Dict, Iterable, List, Optional
from openpyxl import load_workbook
from ofxstatement.parser import StatementParser
from ofxstatement.statement import generate_transaction_id
from ofxstatement.statement import StatementLine


class ExcelStatementParser(StatementParser[List[Any]]):
    """Generic Excel statement parser"""

    mappings: Dict[str, int] = {}
    HEADER_ROW_INDEX: int = 0

    def __init__(self, filename: str) -> None:
        super().__init__()
        self.filename = filename
        self.sheet = self._load_workbook()

    def _load_workbook(self):
        sheet = load_workbook(filename=self.filename, read_only=True).active
        if not sheet:
            raise ValueError("Could not read Excel file")

        # FIX: newer files exported from Handelsbanken returns an incorrect XML
        # range "A1:A1" in their metadata. Resetting dimensions forces openpyxl
        # to scan the actual file content to find the data boundary
        sheet.reset_dimensions()
        return sheet

    def get_cell_value(self, cell: str) -> Optional[Any]:
        return self.sheet[cell].value

    def split_records(self) -> Iterable[List[Any]]:
        return self.sheet.iter_rows(min_row=self.HEADER_ROW_INDEX + 1, values_only=True)

    def parse_record(self, line: List[Any]) -> Optional[StatementLine]:
        stmt_line = StatementLine()
        for field, col in self.mappings.items():
            if col >= len(line):
                raise ValueError(
                    f"Cannot find column {col} in line with {len(line)} items"
                )
            rawvalue = line[col]
            value = self.parse_value(str(rawvalue).strip(), field)
            setattr(stmt_line, field, value)

        if not stmt_line.id:
            stmt_line.id = generate_transaction_id(stmt_line)

        return stmt_line
