from datetime import datetime
from typing import Any, Iterable, List
import re
import itertools
import logging

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import Statement, StatementLine, generate_transaction_id
from openpyxl import load_workbook

class HandelsbankenParser(StatementParser[Any]):
    """Parser for converting Handelsbanken bank statements from Excel to OFX format."""
    
    EXPECTED_HEADERS = [
        "Reskontradatum",
        "Transaktionsdatum",
        "Text",
        "Belopp",
        "Saldo"
    ]
    
    COLUMN_INDICES = {
        'date': 1,
        'description': 2,
        'amount': 3
    }
    
    ACCOUNT_PATTERN = re.compile(r"^(\w+) (\d{3} \d{3} \d{3})$")
    HEADER_ROW_INDEX = 9

    def __init__(self, filename: str) -> None:
        super().__init__()
        self.filename = filename
        self._load_workbook()

    def _load_workbook(self) -> None:
        self.sheet = load_workbook(filename=self.filename, read_only=True).active
        if not self.sheet:
            raise ValueError("Could not read Excel file")

    def _parse_account_info(self) -> tuple[str, str]:
        account_cell = self.sheet["A4"].value
        match = self.ACCOUNT_PATTERN.match(account_cell)
        if not match:
            raise ValueError(f"Invalid account format: {account_cell}")
        return match.group(1), match.group(2)

    def _validate_headers(self, headers: List[Any]) -> None:
        header_values = [cell.value for cell in headers if cell is not None]
        if header_values != self.EXPECTED_HEADERS:
            raise ValueError(
                f"Unexpected column headers. Expected: {self.EXPECTED_HEADERS}, "
                f"Found: {header_values}"
            )

    @staticmethod
    def _parse_date(date_string: str) -> datetime.date:
        try:
            return datetime.strptime(date_string, "%Y-%m-%d").date()
        except ValueError as e:
            raise ValueError(f"Invalid date: {date_string}. Expected format: 'YYYY-MM-DD'.") from e

    def parse(self) -> Statement:
        account_name, account_id = self._parse_account_info()
        self.statement.account_id = account_id
        self.statement.currency = "SEK"
        self.statement.bank_id = "HANDELSBANKEN"
        
        logging.info(f"Parsing statement for account: {account_name} ({account_id})")
        
        rows = self.sheet.iter_rows()
        headers = list(itertools.islice(rows, self.HEADER_ROW_INDEX))[-1]
        self._validate_headers(headers)
        
        self.rows = rows
        return super().parse()

    def split_records(self) -> Iterable[Any]:
        return ([c.value for c in row if c is not None] for row in self.rows)

    def parse_record(self, line: List[Any]) -> StatementLine:
        stmt_line = StatementLine(
            date=self._parse_date(line[self.COLUMN_INDICES['date']]),
            memo=line[self.COLUMN_INDICES['description']],
            amount=line[self.COLUMN_INDICES['amount']],
        )
        stmt_line.id = generate_transaction_id(stmt_line)
        return stmt_line

class HandelsbankenPlugin(Plugin):
    """Plugin for handling Handelsbanken bank statements."""
    
    def get_parser(self, filename: str) -> HandelsbankenParser:
        return HandelsbankenParser(filename)